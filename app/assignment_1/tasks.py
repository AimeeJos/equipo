from __future__ import absolute_import, unicode_literals
from celery import shared_task
import openpyxl
from openpyxl.styles import Font  # PatternFill
import os
from django.conf import settings
from assignment_1.progress import progress_bar


from assignment_1 import extract_data


def write_to_excel_file(self, finaldatas):
    taskid = self.request.id
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    font = Font(bold=True)
    fields = ["Group", "Category", "Code", "Long Description", "Short Description"]

    for i, heading in enumerate(fields):
        cell = worksheet.cell(row=1, column=(i + 1), value=heading)
        cell.font = font

    c = 2
    values = []
    total_steps = len(finaldatas)
    print("=== total: ", total_steps)
    for data in finaldatas:
        values = list(data.values())

        percentage_complete = int(((c - 1) / total_steps) * 100) - 5
        progress_bar(taskid, percentage_complete)

        for i, value in enumerate(values):
            worksheet.cell(row=c, column=(i + 1), value=str(value))
        c = c + 1

    # Create a directory to save the Excel file
    directory = os.path.join(settings.MEDIA_ROOT, "downloads")
    print(f"____directory___{directory} exists?: {os.path.exists(directory)}")
    if not os.path.exists(directory):
        os.makedirs(directory)
    file_path = os.path.join(directory, f"{taskid}.xlsx")
    saved = workbook.save(file_path)
    print(f"file saved: {saved}")

    progress_bar(taskid, 100)


# @shared_task(bind=True)
# def download_data(self, queryset_values):
#     print("______CELERY STARTS_______")
#     # time.sleep(15)

#     task_progress = {
#         "status": 0,
#         "taskid": self.request.id,
#     }
#     channel_layer = get_channel_layer()
#     async_to_sync(channel_layer.group_send)(
#         self.request.id,  # Use the task ID as the channel group name
#         {"type": "task_progress", "percentage_complete": task_progress},
#     )
#     finaldatas = fetch_all_datas(queryset_values)
#     write_to_excel_file(self, finaldatas)
#     return "DONEEE**DOWNLOAD***********"


@shared_task(bind=True)
def do_more_scrapping(self, group, category, url):
    print("======")
    taskid = self.request.id
    progress_bar(taskid, 2)
    datas = extract_data.extract_codes(url, group, category, self.request.id)
    print("**EXTRACTED")
    progress_bar(taskid, 50)
    write_to_excel_file(self, datas)

    return "DONEEE**DOWNLOAD***********"
